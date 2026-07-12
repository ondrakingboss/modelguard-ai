from __future__ import annotations

import hashlib
import re
from collections import Counter
from typing import Any, Iterable

from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from models import AuditResult, Issue, ScoreExplanation, SeverityBreakdown
from parser import ParsedWorkbook, cell_value


FORMULA_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!"}
EXTERNAL_LINK_RE = re.compile(r"\[[^\]]+\.(?:xlsx|xlsm|xlsb|xls)\]", re.IGNORECASE)
CELL_REF_RE = re.compile(r"(?<![A-Z0-9_])\$?([A-Z]{1,3})\$?([1-9][0-9]*)(?![A-Z0-9_])", re.IGNORECASE)
NUMBER_RE = re.compile(r"(?<![A-Z])[-+]?\d+(?:\.\d+)?%?")
ONLY_NUMBERS_OPERATORS_RE = re.compile(r"^=\s*[-+*/().,\s\d%]+$")
TRIVIAL_CONSTANTS = {0.0, 1.0, -1.0, 100.0, -100.0}
REVENUE_RE = re.compile(r"revenue|sales|turnover", re.IGNORECASE)
MARGIN_RE = re.compile(r"margin|gross margin|ebitda margin|operating margin", re.IGNORECASE)
CASH_FLOW_RE = re.compile(r"cash flow|cashflow|cash balance|running cash|ending cash", re.IGNORECASE)
ASSUMPTION_LABEL_RE = re.compile(r"growth|rate|margin|cost|price|inflation|wacc|discount|tax|depreciation", re.IGNORECASE)


def audit_workbook(parsed: ParsedWorkbook) -> AuditResult:
    issues: list[Issue] = []
    workbook = parsed.workbook

    for worksheet in workbook.worksheets:
        issues.extend(_detect_hidden_sheet(worksheet))
        issues.extend(_detect_hidden_rows(worksheet))
        issues.extend(_detect_hidden_columns(worksheet))
        issues.extend(_scan_cells(worksheet))
        issues.extend(_detect_inconsistent_neighboring_formulas(worksheet))
        issues.extend(_detect_assumption_gaps(worksheet))
        issues.extend(_detect_business_anomalies(worksheet))

    # Deduplicate before collapsing row noise
    issues = _dedupe_issues(issues)
    # Detect pairwise circular references (A→B and B→A)
    issues = _detect_pairwise_circulars(workbook, issues)
    # Consolidate related findings on same cell (hardcoded + inconsistent)
    issues = _consolidate_related(issues)
    # Collapse repeated findings across adjacent cells in the same row
    issues = _collapse_row_noise(issues)

    severity_breakdown = Counter(issue.severity for issue in issues)
    breakdown = SeverityBreakdown(
        critical=severity_breakdown["critical"],
        high=severity_breakdown["high"],
        medium=severity_breakdown["medium"],
        low=severity_breakdown["low"],
    )
    score = _score_model(issues, breakdown)
    explanation = _build_score_explanation(score, issues, breakdown)
    summary = _summary_for(score, len(issues), breakdown)

    return AuditResult(
        model_score=score,
        summary=summary,
        issues=issues,
        severity_breakdown=breakdown,
        score_explanation=explanation,
    )


def _detect_hidden_sheet(worksheet: Worksheet) -> list[Issue]:
    if worksheet.sheet_state == "visible":
        return []

    # Check if sheet is empty or has content
    has_content = any(
        cell.value is not None
        for row in worksheet.iter_rows()
        for cell in row
    )

    severity = "high" if has_content else "medium"
    desc_extra = (
        " and contains data or formulas"
        if has_content
        else " but appears empty or metadata-only"
    )

    return [
        _issue(
            severity=severity,
            sheet=worksheet.title,
            cell="",
            category="Hidden Content",
            title="Hidden worksheet detected",
            description=(
                f"Worksheet '{worksheet.title}' is marked as {worksheet.sheet_state}"
                f"{desc_extra}."
            ),
            why_it_matters=(
                "Hidden worksheets can contain assumptions, calculations, or overrides "
                "that materially affect model outputs but are easy to miss during review."
            ),
            suggested_fix=(
                "Unhide the sheet and document whether it is required, archival, "
                "or safe to remove."
            ),
        )
    ]


def _detect_hidden_rows(worksheet: Worksheet) -> list[Issue]:
    issues: list[Issue] = []
    for row_idx, dimension in worksheet.row_dimensions.items():
        if not dimension.hidden:
            continue

        # Check if the row has formulas that could feed visible outputs
        row_cells = list(
            worksheet.iter_rows(min_row=row_idx, max_row=row_idx)
        )
        has_formulas = any(
            isinstance(cell.value, str) and cell.value.startswith("=")
            for row in row_cells
            for cell in row
        )
        has_content = any(
            cell.value is not None
            for row in row_cells
            for cell in row
        )

        if has_formulas:
            severity = "high"
            detail = "and contains formulas that may feed calculations."
        elif has_content:
            severity = "medium"
            detail = "and contains values with no detected downstream impact."
        else:
            severity = "low"
            detail = "but appears empty."

        issues.append(
            _issue(
                severity=severity,
                sheet=worksheet.title,
                cell=str(row_idx),
                category="Hidden Content",
                title="Hidden row detected",
                description=f"Row {row_idx} is hidden {detail}",
                why_it_matters=(
                    "Hidden rows may conceal inputs, manual adjustments, or stale calculations."
                ),
                suggested_fix=(
                    "Unhide the row and either validate its contents or remove it if obsolete."
                ),
            )
        )
    return issues


def _detect_hidden_columns(worksheet: Worksheet) -> list[Issue]:
    issues: list[Issue] = []
    for col_key, dimension in worksheet.column_dimensions.items():
        if not dimension.hidden:
            continue

        # Check if column has numeric content (forecast years, calculations)
        col_index = _column_index(str(col_key))
        col_cells = list(
            worksheet.iter_cols(min_col=col_index, max_col=col_index, values_only=True)
        )
        flat_values = [v for col_tuple in col_cells for v in col_tuple]

        has_numeric = any(
            isinstance(v, (int, float)) for v in flat_values if v is not None
        )
        has_formulas = any(
            isinstance(v, str) and v.startswith("=") for v in flat_values if v is not None
        )

        if has_formulas or has_numeric:
            severity = "high"
            detail = "and contains formulas or numeric data — may hide forecast years or linked calculations."
        else:
            severity = "medium"
            detail = "but appears to contain only labels or metadata."

        issues.append(
            _issue(
                severity=severity,
                sheet=worksheet.title,
                cell=str(col_key),
                category="Hidden Content",
                title="Hidden column detected",
                description=f"Column {col_key} is hidden {detail}",
                why_it_matters=(
                    "Hidden columns can obscure linked calculations, helper inputs, "
                    "or manual overrides."
                ),
                suggested_fix=(
                    "Unhide the column and confirm whether its values are necessary and accurate."
                ),
            )
        )
    return issues


def _scan_cells(worksheet: Worksheet) -> list[Issue]:
    issues: list[Issue] = []
    merged_cells = _merged_cells(worksheet)

    for row in worksheet.iter_rows():
        for cell in row:
            value = cell.value
            text = cell_value(value)

            # Check for error tokens in the value (both bare errors and embedded in formulas)
            has_error = text in FORMULA_ERRORS or any(
                err in text for err in FORMULA_ERRORS
            )
            if has_error:
                issues.append(_formula_error_issue(worksheet, cell, text))

            if _is_formula(value):
                formula = str(value)
                if _has_hardcoded_constants(formula):
                    issues.append(_hardcoded_formula_issue(worksheet, cell, formula))

                if EXTERNAL_LINK_RE.search(formula):
                    issues.append(_external_link_issue(worksheet, cell, formula))

                if _has_self_reference(cell, formula):
                    issues.append(_circular_reference_issue(worksheet, cell))

            if cell.coordinate in merged_cells and text in FORMULA_ERRORS:
                issues.append(_formula_error_issue(worksheet, cell, text))

    return issues


def _detect_inconsistent_neighboring_formulas(worksheet: Worksheet) -> list[Issue]:
    issues: list[Issue] = []

    for row in worksheet.iter_rows():
        formula_cells = [cell for cell in row if _is_formula(cell.value)]
        if len(formula_cells) < 3:
            continue

        patterns = [_formula_pattern(str(cell.value), cell.row, cell.column) for cell in formula_cells]
        counts = Counter(patterns)
        common_pattern, common_count = counts.most_common(1)[0]
        if common_count < 2:
            continue

        for cell, pattern in zip(formula_cells, patterns):
            if pattern != common_pattern:
                issues.append(_inconsistent_formula_issue(worksheet, cell, "row"))

    for col in worksheet.iter_cols():
        formula_cells = [cell for cell in col if _is_formula(cell.value)]
        if len(formula_cells) < 3:
            continue

        patterns = [_formula_pattern(str(cell.value), cell.row, cell.column) for cell in formula_cells]
        counts = Counter(patterns)
        common_pattern, common_count = counts.most_common(1)[0]
        if common_count < 2:
            continue

        for cell, pattern in zip(formula_cells, patterns):
            if pattern != common_pattern:
                issues.append(_inconsistent_formula_issue(worksheet, cell, "column"))

    return _dedupe_issues(issues)


def _detect_assumption_gaps(worksheet: Worksheet) -> list[Issue]:
    if "assumption" not in worksheet.title.lower():
        return []

    issues: list[Issue] = []
    max_row = max(worksheet.max_row, 1)
    max_column = max(worksheet.max_column, 1)

    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
        label = cell_value(row[0].value) if row else ""
        row_values = [cell_value(cell.value) for cell in row]

        if label and all(value == "" for value in row_values[1:]):
            issues.append(
                _issue(
                    severity="medium",
                    sheet=worksheet.title,
                    cell=row[0].coordinate,
                    category="Assumptions",
                    title="Empty assumption line",
                    description=f"Assumption row '{label}' has no entered values.",
                    why_it_matters="Incomplete assumptions can cause downstream formulas to rely on blanks, defaults, or unintended zero values.",
                    suggested_fix="Populate the missing assumption values or remove the unused row.",
                )
            )

    non_empty = sum(1 for row in worksheet.iter_rows() for cell in row if cell_value(cell.value))
    if non_empty == 0:
        issues.append(
            _issue(
                severity="high",
                sheet=worksheet.title,
                cell="A1",
                category="Assumptions",
                title="Empty assumptions sheet",
                description="The assumptions worksheet is empty.",
                why_it_matters="A model with a blank assumptions section may be missing the documented inputs needed to validate outputs.",
                suggested_fix="Add the key operating, financing, and market assumptions used by the model.",
            )
        )

    return issues


def _detect_business_anomalies(worksheet: Worksheet) -> list[Issue]:
    issues: list[Issue] = []
    for row in worksheet.iter_rows():
        label = " ".join(cell_value(cell.value) for cell in row[:3])
        numeric_cells = [(cell, _to_number(cell.value)) for cell in row]
        numeric_cells = [(cell, value) for cell, value in numeric_cells if value is not None]
        if len(numeric_cells) < 2:
            continue

        if REVENUE_RE.search(label):
            issues.extend(_growth_issues(worksheet, numeric_cells))

        if MARGIN_RE.search(label):
            issues.extend(_margin_jump_issues(worksheet, numeric_cells))

        if CASH_FLOW_RE.search(label):
            issues.extend(_negative_cash_flow_issues(worksheet, numeric_cells))

        if ASSUMPTION_LABEL_RE.search(label):
            issues.extend(_assumption_mismatch_issues(worksheet, numeric_cells, label))

    return issues


def _growth_issues(worksheet: Worksheet, numeric_cells: list[tuple[Cell, float]]) -> list[Issue]:
    issues: list[Issue] = []
    for (previous_cell, previous), (current_cell, current) in zip(numeric_cells, numeric_cells[1:]):
        if previous <= 0:
            continue
        growth = (current - previous) / abs(previous)
        flagged = False
        if growth > 0.50:
            desc = f"Revenue grows by {growth:.0%} from {previous_cell.coordinate} to {current_cell.coordinate}."
            flagged = True
        elif growth < -0.30:
            desc = f"Revenue declines by {growth:.0%} from {previous_cell.coordinate} to {current_cell.coordinate}."
            flagged = True
        else:
            continue

        issues.append(
            _issue(
                severity="high",
                sheet=worksheet.title,
                cell=current_cell.coordinate,
                category="Business Logic",
                title="Suspicious revenue growth",
                description=desc,
                why_it_matters=(
                    "Period-over-period revenue changes exceeding 50% growth or -30% decline "
                    "are rare outside distressed or hyper-growth scenarios. Such jumps often "
                    "indicate a broken link, pasted override, or unsupported forecast assumption."
                ),
                suggested_fix="Trace the revenue driver and document the basis for the step change.",
            )
        )
    return issues


def _margin_jump_issues(worksheet: Worksheet, numeric_cells: list[tuple[Cell, float]]) -> list[Issue]:
    issues: list[Issue] = []
    for (previous_cell, previous), (current_cell, current) in zip(numeric_cells, numeric_cells[1:]):
        previous_margin = _normalize_margin(previous)
        current_margin = _normalize_margin(current)
        if previous_margin is None or current_margin is None:
            continue
        jump = abs(current_margin - previous_margin)
        if jump > 0.20:
            issues.append(
                _issue(
                    severity="medium",
                    sheet=worksheet.title,
                    cell=current_cell.coordinate,
                    category="Business Logic",
                    title="Suspicious margin jump",
                    description=f"Margin changes by {jump:.0%} from {previous_cell.coordinate} to {current_cell.coordinate}.",
                    why_it_matters="Large margin changes can materially distort profitability and valuation outputs.",
                    suggested_fix="Validate the margin formula and add support for the operational change if intentional.",
                )
            )
    return issues


def _negative_cash_flow_issues(worksheet: Worksheet, numeric_cells: list[tuple[Cell, float]]) -> list[Issue]:
    issues: list[Issue] = []
    running_total = 0.0
    for cell, value in numeric_cells:
        running_total += value
        if running_total < 0:
            issues.append(
                _issue(
                    severity="high",
                    sheet=worksheet.title,
                    cell=cell.coordinate,
                    category="Business Logic",
                    title="Negative running cash flow",
                    description=f"Running cash flow turns negative at {cell.coordinate}.",
                    why_it_matters="Negative cash flow may imply a funding gap, liquidity issue, or missing financing assumption.",
                    suggested_fix="Review cash inflows, outflows, and financing assumptions for this period.",
                )
            )
            break
    return issues


def _assumption_mismatch_issues(
    worksheet: Worksheet,
    numeric_cells: list[tuple[Cell, float]],
    label: str,
) -> list[Issue]:
    """Flag assumption rows where period-over-period changes suggest stale or mismatched inputs."""
    issues: list[Issue] = []
    values = [v for _, v in numeric_cells]
    if len(values) < 3 or all(v == values[0] for v in values):
        return issues

    # Check for suspiciously static values across periods
    unique_values = len(set(round(v, 6) for v in values))
    if unique_values == 1:
        return issues

    # Flag if a percentage-labeled row has values that look like they should vary but don't
    for i in range(1, len(values)):
        change = abs(values[i] - values[i - 1])
        if values[i - 1] != 0 and change / abs(values[i - 1]) > 1.0:
            issues.append(
                _issue(
                    severity="medium",
                    sheet=worksheet.title,
                    cell=numeric_cells[i][0].coordinate,
                    category="Business Logic",
                    title="Assumption value jump",
                    description=(
                        f"'{label}' changes abruptly at {numeric_cells[i][0].coordinate}. "
                        f"Previous value: {values[i-1]:.4f}, current: {values[i]:.4f}."
                    ),
                    why_it_matters=(
                        "Assumption values that change dramatically between periods may indicate "
                        "a copy-paste error, broken link, or undocumented change in methodology."
                    ),
                    suggested_fix=(
                        "Verify whether the change is intentional. If so, document the rationale. "
                        "If not, restore the consistent assumption series."
                    ),
                )
            )
            break
    return issues


def _has_hardcoded_constants(formula: str) -> bool:
    """Return True if formula contains non-trivial hardcoded numbers.

    Trivial constants (0, 1, -1, 100, -100) and standard modeling
    patterns (percentage conversion, growth formulas) are ignored.
    """
    if ONLY_NUMBERS_OPERATORS_RE.match(formula):
        numbers = [float(n) for n in NUMBER_RE.findall(formula)]
        return any(n not in TRIVIAL_CONSTANTS for n in numbers)

    without_references = CELL_REF_RE.sub("", formula)
    numbers = NUMBER_RE.findall(without_references)
    if not numbers:
        return False

    for num_str in numbers:
        try:
            num = float(num_str.rstrip("%"))
        except ValueError:
            continue
        if num not in TRIVIAL_CONSTANTS:
            return True
    return False


def _detect_pairwise_circulars(workbook: Any, issues: list[Issue]) -> list[Issue]:
    """Detect A→B and B→A circular reference pairs that cross-reference between cells."""
    # Build a map: cell coordinate → set of referenced cell coordinates
    refs: dict[tuple[str, str], set[tuple[str, str]]] = {}

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if not _is_formula(value):
                    continue
                formula = str(value)
                source = (worksheet.title, cell.coordinate)
                refs.setdefault(source, set())
                for match in CELL_REF_RE.finditer(formula):
                    target_coord = f"{match.group(1).upper()}{match.group(2)}"
                    # Assume same sheet for simple references
                    refs[source].add((worksheet.title, target_coord))

    # Find pairs where A references B and B references A
    found: set[tuple[str, str]] = set()
    for (sheet_a, coord_a), targets_a in list(refs.items()):
        for (sheet_b, coord_b) in targets_a:
            if (sheet_b, coord_b) not in refs:
                continue
            if (sheet_a, coord_a) in refs[(sheet_b, coord_b)]:
                # Found a pair — flag both if not already flagged
                pair_key = tuple(sorted([(sheet_a, coord_a), (sheet_b, coord_b)]))
                if pair_key in found:
                    continue
                found.add(pair_key)

                # Determine which cell to flag (avoid double-flagging self-refs already caught)
                for sheet, coord in [(sheet_a, coord_a), (sheet_b, coord_b)]:
                    # Find the OPPOSITE cell in the pair
                    other_sheet = sheet_b if (sheet, coord) == (sheet_a, coord_a) else sheet_a
                    other_coord = coord_b if (sheet, coord) == (sheet_a, coord_a) else coord_a
                    worksheet = workbook[sheet]
                    cell = worksheet[coord]
                    issues.append(
                        _issue(
                            severity="high",
                            sheet=sheet,
                            cell=coord,
                            category="Circular Reference",
                            title=f"Circular reference pair: {coord} ↔ {other_coord}",
                            description=(
                                f"Cell {coord} references {pair_key[1][1] if pair_key[1][0] == sheet else pair_key[0][1]} "
                                f"which references back, forming a circular dependency loop."
                            ),
                            why_it_matters=(
                                "Cross-cell circular references produce unstable outputs that depend on iteration order "
                                "and can mask calculation errors. They often indicate a modeling logic flaw."
                            ),
                            suggested_fix=(
                                "Break the cycle by moving one side of the dependency to a separate schedule "
                                "or using a controlled iterative structure with explicit convergence logic."
                            ),
                        )
                    )

    return issues

def _formula_pattern(formula: str, row: int, column: int) -> str:
    def replace_ref(match: re.Match[str]) -> str:
        ref_col = _column_index(match.group(1))
        ref_row = int(match.group(2))
        return f"R[{ref_row - row}]C[{ref_col - column}]"

    normalized = CELL_REF_RE.sub(replace_ref, formula.upper())
    normalized = re.sub(r"\d+(?:\.\d+)?", "N", normalized)
    return normalized


def _has_self_reference(cell: Cell, formula: str) -> bool:
    cell_ref = f"{get_column_letter(cell.column)}{cell.row}".upper()
    for match in CELL_REF_RE.finditer(formula.upper()):
        if f"{match.group(1).upper()}{match.group(2)}" == cell_ref:
            return True
    return False


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _to_number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip().replace(",", "")
        if stripped.endswith("%"):
            stripped = stripped[:-1]
            try:
                return float(stripped) / 100
            except ValueError:
                return None
        try:
            return float(stripped)
        except ValueError:
            return None
    return None


def _normalize_margin(value: float) -> float | None:
    if -1.0 <= value <= 1.0:
        return value
    if -100.0 <= value <= 100.0:
        return value / 100
    return None


def _column_index(column_letters: str) -> int:
    value = 0
    for char in column_letters.upper().replace("$", ""):
        value = value * 26 + ord(char) - ord("A") + 1
    return value


def _merged_cells(worksheet: Worksheet) -> set[str]:
    coordinates: set[str] = set()
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                coordinates.add(f"{get_column_letter(column)}{row}")
    return coordinates


def _dedupe_issues(issues: Iterable[Issue]) -> list[Issue]:
    seen: set[tuple[str, str, str, str]] = set()
    unique: list[Issue] = []
    for issue in issues:
        key = (issue.sheet, issue.cell, issue.category, issue.title)
        if key in seen:
            continue
        seen.add(key)
        unique.append(issue)
    return unique


def _collapse_row_noise(issues: list[Issue]) -> list[Issue]:
    """Group repeated findings across adjacent cells in the same row.

    If the same rule fires on adjacent cells in the same row for the same
    category and title, collapse into one grouped finding with a cell range.
    Individual cells are preserved in the description.
    """
    if len(issues) < 3:
        return issues

    # Group by (sheet, category, title) — find consecutive cell patterns
    from collections import defaultdict
    import re as _re

    def _normalize_title(t: str) -> str:
        """Strip cell references from titles so patterns group correctly."""
        return _re.sub(r"at [A-Z]+\d+", "at XX", t)

    groups: dict[tuple[str, str, str], list[Issue]] = defaultdict(list)
    for issue in issues:
        if not issue.cell:
            continue
        key = (issue.sheet, issue.category, _normalize_title(issue.title))
        groups[key].append(issue)

    collapsed: dict[str, Issue] = {}
    removed_ids: set[str] = set()

    for key, grouped in groups.items():
        if len(grouped) < 3:
            continue

        # Extract cell rows — group by row
        from collections import defaultdict
        row_groups: dict[int, list[Issue]] = defaultdict(list)
        for issue in grouped:
            row_num = _extract_row(issue.cell)
            if row_num:
                row_groups[row_num].append(issue)

        for row_num, row_issues in row_groups.items():
            if len(row_issues) < 3:
                continue

            # Sort by column
            sorted_issues = sorted(row_issues, key=lambda i: _extract_col(i.cell) or 0)

            # Find runs of consecutive columns
            runs: list[list[Issue]] = []
            current_run: list[Issue] = [sorted_issues[0]]
            for i in range(1, len(sorted_issues)):
                prev_col = _extract_col(sorted_issues[i - 1].cell)
                curr_col = _extract_col(sorted_issues[i].cell)
                if prev_col is not None and curr_col is not None and curr_col == prev_col + 1:
                    current_run.append(sorted_issues[i])
                else:
                    if len(current_run) >= 3:
                        runs.append(current_run)
                    current_run = [sorted_issues[i]]
            if len(current_run) >= 3:
                runs.append(current_run)

            for run in runs:
                first = run[0]
                last = run[-1]
                cells = sorted([i.cell for i in run], key=lambda c: (_extract_col(c) or 0))
                cell_range = f"{first.sheet}!{cells[0]}:{cells[-1]}"
                affected_list = ", ".join(cells[:8])
                if len(cells) > 8:
                    affected_list += f", ... ({len(cells)} cells total)"

                collapsed_id = f"{first.id}_grouped_{cell_range}"
                grouped_issue = Issue(
                    id=collapsed_id,
                    severity=first.severity,
                    sheet=first.sheet,
                    cell=cell_range,
                    category=first.category,
                    title=f"{first.title} (×{len(cells)} cells)",
                    description=(
                        f"{first.description} Pattern repeats across {len(cells)} "
                        f"adjacent cells in row {row_num}: {affected_list}."
                    ),
                    why_it_matters=first.why_it_matters,
                    suggested_fix=first.suggested_fix,
                )
                collapsed[collapsed_id] = grouped_issue
                removed_ids.update(i.id for i in run)

    if not collapsed:
        return issues

    result = [i for i in issues if i.id not in removed_ids]
    result.extend(collapsed.values())
    return result


def _extract_row(cell_ref: str) -> int | None:
    """Extract row number from a cell reference like 'B7' or 'A1'."""
    import re
    m = re.search(r"(\d+)", cell_ref)
    return int(m.group(1)) if m else None


def _extract_col(cell_ref: str) -> int | None:
    """Extract column index from a cell reference like 'B7'."""
    import re
    m = re.match(r"([A-Z]+)", cell_ref)
    if not m:
        return None
    return _column_index(m.group(1))


def _formula_error_issue(worksheet: Worksheet, cell: Cell, error: str) -> Issue:
    return _issue(
        severity="critical",
        sheet=worksheet.title,
        cell=cell.coordinate,
        category="Formula Error",
        title=f"Formula error {error}",
        description=f"Cell {cell.coordinate} contains Excel error value {error}.",
        why_it_matters="Formula errors can cascade through linked schedules and make outputs unreliable.",
        suggested_fix="Trace precedent cells, correct the broken formula, and verify dependent calculations.",
    )


def _hardcoded_formula_issue(worksheet: Worksheet, cell: Cell, formula: str) -> Issue:
    return _issue(
        severity="medium",
        sheet=worksheet.title,
        cell=cell.coordinate,
        category="Formula Integrity",
        title="Hardcoded constant inside formula",
        description=f"Formula '{formula}' contains hardcoded numeric input.",
        why_it_matters="Hardcoded constants make assumptions difficult to audit and can bypass the model's input controls.",
        suggested_fix="Move the constant to a clearly labeled assumption cell and reference that cell in the formula.",
    )


def _external_link_issue(worksheet: Worksheet, cell: Cell, formula: str) -> Issue:
    return _issue(
        severity="high",
        sheet=worksheet.title,
        cell=cell.coordinate,
        category="External Links",
        title="External workbook link",
        description=f"Formula '{formula}' references another workbook.",
        why_it_matters="External workbook links can break, pull stale data, or make the model impossible to reproduce for reviewers.",
        suggested_fix="Import the required source data into the model or document and package the linked workbook.",
    )


def _circular_reference_issue(worksheet: Worksheet, cell: Cell) -> Issue:
    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
    formula = str(cell.value) if cell.value else ""
    return _issue(
        severity="high",
        sheet=worksheet.title,
        cell=cell.coordinate,
        category="Circular Reference",
        title=f"Circular reference at {cell_ref}",
        description=(
            f"Cell {cell_ref} directly references itself in formula: {formula}. "
            "This creates a dependency loop that Excel resolves through iteration, "
            "producing unstable or iteration-dependent outputs."
        ),
        why_it_matters=(
            "Circular references can produce different results on recalc, mask calculation "
            "errors, and make the model fragile. They are frequently used to hide modeling "
            "issues or bypass proper schedule design."
        ),
        suggested_fix=(
            "Rewrite the formula to remove the self-reference. Move iterative logic "
            "into a controlled schedule with explicit iteration and break conditions."
        ),
    )


def _inconsistent_formula_issue(worksheet: Worksheet, cell: Cell, axis: str) -> Issue:
    return _issue(
        severity="medium",
        sheet=worksheet.title,
        cell=cell.coordinate,
        category="Formula Integrity",
        title="Inconsistent neighboring formula",
        description=f"Formula in {cell.coordinate} differs from nearby formulas in the same {axis}.",
        why_it_matters="Inconsistent formulas often indicate accidental overwrites, copied formulas with broken references, or hidden manual adjustments.",
        suggested_fix="Compare the formula against adjacent periods or line items and restore the intended pattern.",
    )


def _consolidate_related(issues: list[Issue]) -> list[Issue]:
    """Merge hardcoded constant + inconsistent formula on same cell into one finding."""
    from collections import defaultdict

    # Group by (sheet, cell)
    by_cell: dict[tuple[str, str], list[Issue]] = defaultdict(list)
    for issue in issues:
        by_cell[(issue.sheet, issue.cell)].append(issue)

    merged: list[Issue] = []
    removed: set[str] = set()

    for (sheet, cell), cell_issues in by_cell.items():
        hardcoded = [i for i in cell_issues if "hardcoded" in i.title.lower() or "hardcoded" in i.category.lower()]
        inconsistent = [i for i in cell_issues if "inconsistent" in i.title.lower() or "inconsistent" in i.category.lower()]

        if hardcoded and inconsistent:
            # Merge: keep the inconsistent as base, enrich with hardcoded info
            hc = hardcoded[0]
            inc = inconsistent[0]
            merged_id = f"{hc.id}_consolidated"
            merged.append(
                Issue(
                    id=merged_id,
                    severity=_worse_severity(hc.severity, inc.severity),
                    sheet=sheet,
                    cell=cell,
                    category="Formula Integrity",
                    title="Inconsistent formula with embedded hardcoded assumption",
                    description=(
                        f"{inc.description} Additionally: {hc.description}"
                    ),
                    why_it_matters=(
                        "This cell both breaks the formula pattern AND contains a hardcoded numeric "
                        "assumption. Together these are a strong signal of a manual override or model error."
                    ),
                    suggested_fix=(
                        "Restore the consistent formula pattern and move any required constants "
                        "to a labeled assumptions cell."
                    ),
                )
            )
            removed.update(i.id for i in hardcoded)
            removed.update(i.id for i in inconsistent)
        elif hardcoded or inconsistent:
            # If only one of the two, prefer the more specific one
            kept = hardcoded[0] if hardcoded else inconsistent[0]
            if kept.id not in removed:
                merged.append(kept)

    # Keep all non-merged issues
    for issue in issues:
        if issue.id not in removed:
            dup_check = (issue.sheet, issue.cell)
            if dup_check not in by_cell or not (
                any(i.id in removed for i in by_cell[dup_check] if i.id == issue.id)
            ):
                if issue.id not in {m.id for m in merged}:
                    merged.append(issue)

    return merged


def _worse_severity(a: str, b: str) -> str:
    order = {"low": 0, "medium": 1, "high": 2, "critical": 3}
    return a if order.get(a, 0) >= order.get(b, 0) else b


def _score_model(issues: list[Issue], breakdown: SeverityBreakdown) -> int:
    """Calibrated scoring with per-severity caps, category caps, and floor rules.

    Base score: 100. Penalties reduce it. Floor rules set maximum ceilings.
    """
    # ── Penalty calculation ──
    critical_count = breakdown.critical
    high_count = breakdown.high
    medium_count = breakdown.medium
    low_count = breakdown.low

    # Count formula-integrity category findings for its own cap
    formula_noise_count = sum(1 for i in issues if i.category == "Formula Integrity")
    # Count row-grouped findings (title contains ×N cells)
    row_group_count = sum(1 for i in issues if "×" in i.title and "cells" in i.title)

    # Per-severity penalties (with caps)
    critical_penalty = min(critical_count * 25, 60)
    high_penalty = min(high_count * 12, 45)
    medium_penalty = min(medium_count * 4, 28)
    low_penalty = min(low_count * 1, 8)

    # Category caps
    formula_noise_penalty = min(formula_noise_count * 3, 20)
    # Row groups: count each group once (already counted in severity), but cap total
    row_group_excess = max(0, row_group_count - 5) * 1

    total_penalty = (
        critical_penalty
        + high_penalty
        + medium_penalty
        + low_penalty
        + row_group_excess
    )

    score = max(0, 100 - total_penalty)

    # ── Floor rules ──
    has_critical = critical_count > 0
    has_formula_error = any(
        i.category == "Formula Error" and i.severity == "critical" for i in issues
    )
    has_circular = any(i.category == "Circular Reference" for i in issues)
    high_count_3plus = high_count >= 3
    hidden_feeds_calc = any(
        i.category == "Hidden Content" and i.severity in ("high", "critical")
        for i in issues
    )

    if has_formula_error and has_circular:
        score = min(score, 45)
    elif has_circular:
        score = min(score, 60)
    elif has_critical:
        score = min(score, 75)
    elif high_count_3plus:
        score = min(score, 55)
    elif hidden_feeds_calc:
        score = min(score, 55)

    return max(0, min(100, score))


def _build_score_explanation(
    score: int, issues: list[Issue], breakdown: SeverityBreakdown
) -> ScoreExplanation:
    """Build a human-readable score explanation."""
    if score >= 80:
        band = "Healthy"
    elif score >= 60:
        band = "Mostly Sound"
    elif score >= 40:
        band = "Moderate Risk"
    elif score >= 20:
        band = "High Risk"
    else:
        band = "Critical Risk"

    # Main drivers
    drivers: list[str] = []
    if breakdown.critical > 0:
        drivers.append(f"{breakdown.critical} critical finding(s) — formula errors or data corruption")
    if any(i.category == "Circular Reference" for i in issues):
        drivers.append("Circular reference(s) detected — iteration-dependent output risk")
    if breakdown.high > 0:
        drivers.append(f"{breakdown.high} high-severity finding(s) — hidden content, business logic anomalies")
    if breakdown.medium > 0:
        drivers.append(f"{breakdown.medium} medium-severity finding(s) — formula integrity, assumptions")
    if not drivers:
        drivers.append("No significant risks detected")

    # Penalty breakdown
    penalties: dict[str, int] = {
        "critical_weight": -25,
        "high_weight": -12,
        "medium_weight": -4,
        "low_weight": -1,
        "critical_total": min(breakdown.critical * 25, 60),
        "high_total": min(breakdown.high * 12, 45),
        "medium_total": min(breakdown.medium * 4, 28),
        "low_total": min(breakdown.low * 1, 8),
    }

    # Caps applied
    caps: list[str] = []
    if breakdown.medium * 4 > 28:
        caps.append(f"Medium penalty capped at -28 (actual: -{breakdown.medium * 4})")
    if breakdown.high * 12 > 45:
        caps.append(f"High penalty capped at -45 (actual: -{breakdown.high * 12})")
    if breakdown.critical * 25 > 60:
        caps.append(f"Critical penalty capped at -60 (actual: -{breakdown.critical * 25})")
    formula_count = sum(1 for i in issues if i.category == "Formula Integrity")
    if formula_count * 3 > 20:
        caps.append(f"Formula noise cap applied: {formula_count} findings limited to -20")

    # Floor rules applied
    floors: list[str] = []
    has_formula_error = any(i.category == "Formula Error" and i.severity == "critical" for i in issues)
    has_circular = any(i.category == "Circular Reference" for i in issues)
    if has_formula_error and has_circular:
        floors.append("Critical formula error + circular reference → max score 45")
    elif has_circular:
        floors.append("Circular reference → max score 60")
    elif breakdown.critical > 0:
        floors.append("Critical findings → max score 75")
    elif breakdown.high >= 3:
        floors.append("3+ high findings → max score 55")

    # Narrative
    why_not_higher = _build_why_not_higher(issues, breakdown, caps, floors)
    why_not_lower = _build_why_not_lower(issues, breakdown, caps)

    return ScoreExplanation(
        score=score,
        score_band=band,
        main_drivers=drivers,
        penalty_breakdown=penalties,
        caps_applied=caps,
        floor_rules_applied=floors,
        why_not_lower=why_not_lower,
        why_not_higher=why_not_higher,
    )


def _build_why_not_higher(
    issues: list[Issue],
    breakdown: SeverityBreakdown,
    caps: list[str],
    floors: list[str],
) -> str:
    parts: list[str] = []
    if breakdown.critical > 0:
        parts.append(
            f"{breakdown.critical} critical formula error(s) exist and must be resolved "
            f"before the model can be considered reliable"
        )
    if any(i.category == "Circular Reference" for i in issues):
        parts.append(
            "circular references create unstable outputs that no amount of other quality "
            "can compensate for"
        )
    if breakdown.high >= 3:
        parts.append(
            f"{breakdown.high} high-severity findings (hidden content, business logic) "
            f"indicate structural model issues"
        )
    if breakdown.medium > 5:
        parts.append(
            f"{breakdown.medium} medium findings — while individually minor, the volume "
            f"suggests systematic formula discipline issues"
        )
    if not parts:
        parts.append("minor formula integrity findings prevent a perfect score")
    if caps:
        parts.append("penalty caps limited further score erosion")
    return ". ".join(parts) + "."


def _build_why_not_lower(
    issues: list[Issue],
    breakdown: SeverityBreakdown,
    caps: list[str],
) -> str:
    parts: list[str] = []
    if breakdown.critical == 0:
        parts.append("no critical formula errors (REF, DIV/0, VALUE)")
    if not any(i.category == "Circular Reference" for i in issues):
        parts.append("no circular references detected")
    if breakdown.medium < 10:
        parts.append("medium findings are below the noise floor threshold")
    else:
        parts.append(
            f"medium findings were capped ({len(caps)} caps applied) to prevent "
            f"repeated row-level noise from dominating the score"
        )
    if breakdown.high < 3:
        parts.append("fewer than 3 high-severity issues — no structural risk indicator")
    if not parts:
        parts.append("penalty caps prevented score collapse")
    return ". ".join(parts) + "."


def _summary_for(score: int, issue_count: int, breakdown: SeverityBreakdown) -> str:
    if issue_count == 0:
        return "No audit issues were detected. The workbook appears structurally clean based on the configured checks."
    if score >= 80:
        posture = "mostly sound"
    elif score >= 60:
        posture = "moderate risk"
    elif score >= 40:
        posture = "high risk"
    else:
        posture = "critical risk"

    return (
        f"Detected {issue_count} issue(s). Model risk is {posture}, with "
        f"{breakdown.critical} critical, {breakdown.high} high, "
        f"{breakdown.medium} medium, and {breakdown.low} low severity finding(s)."
    )


def _issue(
    *,
    severity: str,
    sheet: str,
    cell: str,
    category: str,
    title: str,
    description: str,
    why_it_matters: str,
    suggested_fix: str,
) -> Issue:
    digest = hashlib.sha1(
        "|".join([severity, sheet, cell, category, title, description]).encode("utf-8")
    ).hexdigest()[:10]
    return Issue(
        id=f"issue_{digest}",
        severity=severity,  # type: ignore[arg-type]
        sheet=sheet,
        cell=cell,
        category=category,
        title=title,
        description=description,
        why_it_matters=why_it_matters,
        suggested_fix=suggested_fix,
    )
