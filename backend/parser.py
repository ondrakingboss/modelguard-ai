from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


@dataclass(frozen=True)
class ParsedWorkbook:
    path: Path
    workbook: Workbook
    dataframes: dict[str, pd.DataFrame]


def parse_excel_file(path: str | Path) -> ParsedWorkbook:
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, data_only=False)
    dataframes = _load_dataframes(workbook_path)
    return ParsedWorkbook(path=workbook_path, workbook=workbook, dataframes=dataframes)


def _load_dataframes(path: Path) -> dict[str, pd.DataFrame]:
    try:
        sheets: dict[str, pd.DataFrame] = pd.read_excel(
            path,
            sheet_name=None,
            header=None,
            engine="openpyxl",
        )
    except Exception:
        return {}

    return {name: frame.where(pd.notna(frame), None) for name, frame in sheets.items()}


def cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
