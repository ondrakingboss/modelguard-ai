import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# === Sheet 1: Forecast (with intentional errors) ===
ws1 = wb.active
ws1.title = "Forecast"

# Header styling
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
thin_border = Border(
    left=Side(style='thin', color='333333'),
    right=Side(style='thin', color='333333'),
    top=Side(style='thin', color='333333'),
    bottom=Side(style='thin', color='333333'),
)
number_fmt = '#,##0'

# Revenue section
headers = ['', 'Q1 2025', 'Q2 2025', 'Q3 2025', 'Q4 2025', 'Q1 2026', 'Q2 2026', 'Q3 2026', 'Q4 2026']
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# Row labels
rows = [
    ('Product Revenue', [250000, 280000, 310000, 350000, 390000, 450000, 520000, 610000]),
    ('Service Revenue', [180000, 195000, 210000, 230000, 250000, 280000, 310000, 350000]),
    ('Total Revenue', [], [], [], [], [], [], [], []),  # formulas
    ('COGS', [130000, 140000, 155000, 175000, 195000, 220000, 260000, 305000]),
    ('Gross Profit', [], [], [], [], [], [], [], []),  # formulas
    ('Gross Margin %', [], [], [], [], [], [], [], []),  # formulas
    ('OpEx - Marketing', [45000, 48000, 52000, 58000, 62000, 70000, 80000, 95000]),
    ('OpEx - R&D', [30000, 32000, 35000, 38000, 42000, 48000, 55000, 65000]),
    ('OpEx - G&A', [25000, 26000, 28000, 30000, 32000, 35000, 38000, 42000]),
    ('Total OpEx', [], [], [], [], [], [], [], []),  # formulas
    ('EBITDA', [], [], [], [], [], [], [], []),  # formulas
    ('EBITDA Margin %', [], [], [], [], [], [], [], []),  # formulas
]

for row_idx, (label, *values) in enumerate(rows, 2):
    ws1.cell(row=row_idx, column=1, value=label)
    if isinstance(values[0], list):
        for col_idx, val in enumerate(values[0], 2):
            ws1.cell(row=row_idx, column=col_idx, value=val)
            ws1.cell(row=row_idx, column=col_idx).number_format = number_fmt

# Add formulas with some INTENTIONAL ERRORS
# Total Revenue = Product + Service (correct)
for col in range(2, 10):
    ws1.cell(row=4, column=col, value=f"=B{2}+B{3}").number_format = number_fmt  # Row 4 = correct

# Gross Profit = Total Revenue - COGS
for col in range(2, 10):
    ws1.cell(row=6, column=col, value=f"=B{4}-B{5}").number_format = number_fmt

# Gross Margin %
for col in range(2, 10):
    ws1.cell(row=7, column=col, value=f"=B{6}/B{4}").number_format = '0.0%'

# Total OpEx
for col in range(2, 10):
    ws1.cell(row=11, column=col, value=f"=B{8}+B{9}+B{10}").number_format = number_fmt

# EBITDA = Gross Profit - Total OpEx
for col in range(2, 10):
    ws1.cell(row=12, column=col, value=f"=B{6}-B{11}").number_format = number_fmt

# EBITDA Margin %
for col in range(2, 10):
    ws1.cell(row=13, column=col, value=f"=B{12}/B{4}").number_format = '0.0%'

# INTENTIONAL ERRORS:
# 1. #REF! error in Q3 2026 (column H, row 3 - Service Revenue)
ws1.cell(row=3, column=8, value="=#REF!")  # Simulate broken reference
# 2. Hardcoded number in formula cell (Q4 2026 Product Revenue)
ws1.cell(row=2, column=9, value=610000)  # Should be a formula or reference — hardcoded
# 3. Inconsistent formula — Q2 2026 Marketing uses different formula than neighbors
ws1.cell(row=8, column=7, value=70000)  # Hardcoded instead of formula

# 4. Suspicious revenue growth — Q2 2026 (col F) to Q3 2026 (col G) 
#    Product: 450000 to 520000 = ~16% growth — not flagged
#    Let's make a big jump: add a row with 137% growth
ws1.insert_rows(8)
ws1.cell(row=8, column=1, value="Growth Rate (hidden row)")
ws1.row_dimensions[8].hidden = True  # HIDDEN ROW
for col in range(2, 10):
    ws1.cell(row=8, column=col, value=0)

# === Sheet 2: Assumptions (with issues) ===
ws2 = wb.create_sheet("Assumptions")

assumptions = [
    ('Growth Rate', '15%', '15%', '15%', '15%', '15%', '15%', '15%', '15%'),
    ('Churn Rate', '5%', '5%', '4%', '4%', '3%', '3%', '2%', '2%'),
    ('Avg Deal Size', '$45,000', '$48,000', '$52,000', '$55,000', '$58,000', '$62,000', '$65,000', '$70,000'),
    ('Headcount', 24, 28, 32, 36, 40, 45, 50, 55),
    ('Discount Rate', '', '', '', '', '', '', '', ''),  # EMPTY ASSUMPTION
    ('Tax Rate', '21%', '21%', '21%', '21%', '21%', '21%', '21%', '21%'),
]

ws2.cell(row=1, column=1, value="Key Assumptions")
ws2.cell(row=1, column=1).font = Font(bold=True, size=14)

for col, h in enumerate(headers, 1):
    cell = ws2.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill

for row_idx, (label, *values) in enumerate(assumptions, 4):
    ws2.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
    for col_idx, val in enumerate(values[0], 2):
        ws2.cell(row=row_idx, column=col_idx, value=val)

# HIDE a column (Q2 2026 = column F)
ws2.column_dimensions['F'].hidden = True

# Hidden sheet
ws3 = wb.create_sheet("Sensitivities")
ws3.sheet_state = 'hidden'  # HIDDEN SHEET
ws3.cell(row=1, column=1, value="Sensitivity Analysis")
ws3.cell(row=2, column=1, value="Best Case").font = Font(bold=True)
ws3.cell(row=3, column=1, value="Base Case").font = Font(bold=True)
ws3.cell(row=4, column=1, value="Worst Case").font = Font(bold=True)
ws3.cell(row=2, column=2, value=850000)
ws3.cell(row=3, column=2, value=620000)
ws3.cell(row=4, column=2, value=380000)

# External link reference
ws1.cell(row=15, column=1, value="External Ref")
ws1.cell(row=15, column=2, value="=[Board_Plan_2025.xlsx]Sheet1!$A$1")

# Save
path = "/Users/chrisjanov/modelguard-ai/sample_model.xlsx"
wb.save(path)
print(f"Sample model created: {path}")
print("Contains: Forecast sheet, Assumptions sheet, hidden Sensitivities sheet")
print("Errors: #REF!, hardcoded formula, hidden row, hidden column, empty assumption, external link")
